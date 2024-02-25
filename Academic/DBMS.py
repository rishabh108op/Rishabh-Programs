from tkinter import*
from PIL import Image,ImageTk
from tkinter import ttk
import mysql.connector
from tkinter import messagebox
from openpyxl import Workbook

class PharmacyManagementSystem:
    def __init__(self,root):
        self.root=root
        self.root.title("Pharmacy Management System")
        self.root.geometry("1000x800")
        lbltitle=Label(self.root,text="PHARAMACY MANAGEMENT SYSTEM",bg='white',fg="red",font=("times new roman",20,"bold"),padx=2,pady=4)
        lbltitle.pack(side=TOP)   #for the title fill to fill the gap
        
        img1=Image.open("E:\\Gate CN Notes\\28471597.jpg")
        img1=img1.resize((30,30),Image.Resampling.LANCZOS)
        self.photoimg1=ImageTk.PhotoImage(img1)
        b1=Button(self.root,image=self.photoimg1,borderwidth=0)
        b1.place(x=70,y=4)

        img2=Image.open("E:\\Gate CN Notes\\28471597.jpg")
        img2=img2.resize((30,30),Image.Resampling.LANCZOS)
        self.photoimg2=ImageTk.PhotoImage(img2)
        b1=Button(self.root,image=self.photoimg2,borderwidth=0)
        b1.place(x=1180,y=4)
        
        #data frame
        DataFrame=Frame(self.root,bd=15,relief=RIDGE)
        DataFrame.place(x=0,y=40,width=1275,height=600)
        
        
################################################  MedicineInfo Table  #######################################################
        self.refinfo_var=StringVar()
        self.storeid_var=StringVar()
        self.typemed_var=StringVar()
        self.sideeffect_var=StringVar()
        self.price_var=StringVar()

        DataFrameLeft=LabelFrame(DataFrame,bd=10,relief=RIDGE,padx=20,text="Medicine Information",fg="darkgreen",font=("arial",10,"bold"))
        DataFrameLeft.place(x=0,y=5,width=420,height=360)
        
        lblstoreid=Label(DataFrameLeft,font=("arial",12,"bold"),text="Store ID",fg="black",padx=2,pady=6)
        lblstoreid.grid(row=0,column=0,sticky=W)
        ref_combo=ttk.Combobox(DataFrameLeft,textvariable=self.storeid_var,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combo["values"]=("1001","1002","1003","1004","1005")
        ref_combo.current(0)
        ref_combo.grid(row=0,column=1)
        
        lblExDate=Label(DataFrameLeft,font=("arial",12,"bold"),text="Ref",padx=2,pady=6)
        lblExDate.grid(row=2,column=0,sticky=W)
        connref=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursorref=connref.cursor()
        my_cursorref.execute("select Ref from MedicineStock")
        rowsref=my_cursorref.fetchall()
        ref_combowalabox=ttk.Combobox(DataFrameLeft,textvariable=self.refinfo_var,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combowalabox["values"]=rowsref
        ref_combowalabox.current(0)
        ref_combowalabox.grid(row=2,column=1)
        
        lblTypeofMedicine=Label(DataFrameLeft,font=("arial",12,"bold"),text="Type Of Medicine",padx=2,pady=6)
        lblTypeofMedicine.grid(row=1,column=0,sticky=W)
        comTypeofMedicine=ttk.Combobox(DataFrameLeft,textvariable=self.typemed_var,state="readonly",font=("arial",12,"bold"),width=22)
        comTypeofMedicine["values"]=("Tablet","Liquid","Capsules","Topical Medicines","Drops","Inhales","Injection")
        comTypeofMedicine.current(0)
        comTypeofMedicine.grid(row=1,column=1)
        
        lblSideEffects=Label(DataFrameLeft,font=("arial",12,"bold"),text="Side Effect:",padx=2,pady=6)
        lblSideEffects.grid(row=3,column=0,sticky=W)
        txtSideEffects=Entry(DataFrameLeft,textvariable=self.sideeffect_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtSideEffects.grid(row=3,column=1)
        
        lblPrice=Label(DataFrameLeft,font=("arial",12,"bold"),text="Tablets Price:",padx=2,pady=6)
        lblPrice.grid(row=4,column=0,sticky=W)
        txtPrice=Entry(DataFrameLeft,textvariable=self.price_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtPrice.grid(row=4,column=1)
        
        side_leftframe=Frame(DataFrameLeft,bd=4,relief=RIDGE,bg="white")
        side_leftframe.place(x=0,y=180,width=290,height=150)
        
        scc_x=ttk.Scrollbar(side_leftframe,orient=HORIZONTAL)
        scc_x.pack(side=BOTTOM,fill=X)
        scc_y=ttk.Scrollbar(side_leftframe,orient=VERTICAL)
        scc_y.pack(side=RIGHT,fill=Y)
        
        self.medicineinfo_table=ttk.Treeview(side_leftframe,columns=("storeid","ref","typeofmed","sideeffect","pricemed"),xscrollcommand=scc_x.set,yscrollcommand=scc_y.set)
        
        scc_x.config(command=self.medicineinfo_table.xview)
        scc_y.config(command=self.medicineinfo_table.yview)
        self.medicineinfo_table["show"]="headings"
        self.medicineinfo_table.heading("storeid",text="Store ID")
        self.medicineinfo_table.heading("ref",text="Ref")
        self.medicineinfo_table.heading("typeofmed",text="Type Of Med")
        self.medicineinfo_table.heading("sideeffect",text="Side Effect")
        self.medicineinfo_table.heading("pricemed",text="Price")
        self.medicineinfo_table.pack(fill=BOTH,expand=1)

        self.medicineinfo_table.column("storeid",width=80)
        self.medicineinfo_table.column("ref",width=80)
        self.medicineinfo_table.column("typeofmed",width=80)
        self.medicineinfo_table.column("sideeffect",width=80)
        self.medicineinfo_table.column("pricemed",width=80)
        self.fetch_dataMedinfo()
        self.medicineinfo_table.bind("<ButtonRelease-1>",self.Medget_cursorinfo)
        
        #MedicineInfo button frame
        down_frame1=Frame(DataFrameLeft,border=4,relief=RIDGE,bg="darkgreen")
        down_frame1.place(x=300,y=185,width=80,height=140)
        
        btnAddmed=Button(down_frame1,text="ADD",font=("arial",8,"bold"),bg="lime",pady=2,fg="white",width=9,command=self.AddMedinfo)
        btnAddmed.grid(row=0,column=0)
        btnAddmed=Button(down_frame1,text="UPDATE",font=("arial",8,"bold"),bg="purple",pady=2,fg="white",width=9,command=self.UpdateMedinfo)
        btnAddmed.grid(row=1,column=0)
        btnAddmed=Button(down_frame1,text="DELETE",font=("arial",8,"bold"),bg="red",pady=2,fg="white",width=9,command=self.DeleteMedinfo)
        btnAddmed.grid(row=2,column=0)
        btnAddmed=Button(down_frame1,text="CLEAR",font=("arial",8,"bold"),bg="orange",pady=2,fg="white",width=9,command=self.clearmedinfo)
        btnAddmed.grid(row=3,column=0)
        btnAddmed=Button(down_frame1,text="REFRESH",font=("arial",8,"bold"),bg="green",pady=2,fg="white",width=9,command=self.refreshinfo)
        btnAddmed.grid(row=4,column=0)

        ##########################################  PATIENT  tABLE  ##########################################################
        
        
        #Patient Table Variables
        self.patientid_var=StringVar()
        self.patientnamevar=StringVar()
        self.addresspatient_var=StringVar()
        self.patientdoctorid_var=StringVar()
        self.symptoms_var=StringVar()
        
        DataFramemid=LabelFrame(DataFrame,bd=10,relief=RIDGE,padx=20,text="Patient Info",fg="darkgreen",font=("arial",10,"bold"))
        DataFramemid.place(x=420,y=5,width=405,height=360)
        
        lblpatientid=Label(DataFramemid,font=("arial",12,"bold"),text="Patient ID",fg="black",padx=2,pady=6)
        lblpatientid.grid(row=0,column=0,sticky=W)
        txtpatientid=Entry(DataFramemid,textvariable=self.patientid_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtpatientid.grid(row=0,column=1)
        
        lblExDate=Label(DataFramemid,font=("arial",12,"bold"),text="Patient Name",padx=2,pady=6)
        lblExDate.grid(row=1,column=0,sticky=W)
        txtExDate=Entry(DataFramemid,textvariable=self.patientnamevar,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtExDate.grid(row=1,column=1)
        
        lblAddresspatient=Label(DataFramemid,font=("arial",12,"bold"),text="Address",padx=2,pady=6)
        lblAddresspatient.grid(row=2,column=0,sticky=W)
        txtAddresspatient=Entry(DataFramemid,textvariable=self.addresspatient_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtAddresspatient.grid(row=2,column=1)
        
        lblSymtoms=Label(DataFramemid,font=("arial",12,"bold"),text="Symtoms",padx=2,pady=6)
        lblSymtoms.grid(row=3,column=0,sticky=W)
        txtSymtoms=Entry(DataFramemid,textvariable=self.symptoms_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtSymtoms.grid(row=3,column=1)
        
        lblDoctoridpatient=Label(DataFramemid,font=("arial",12,"bold"),text="Doctor ID",padx=2,pady=6)
        lblDoctoridpatient.grid(row=4,column=0,sticky=W)
        txtDoctoridpatient=Entry(DataFramemid,textvariable=self.patientdoctorid_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtDoctoridpatient.grid(row=4,column=1)
        
        side_leftpatientframe=Frame(DataFramemid,bd=4,relief=RIDGE,bg="white")
        side_leftpatientframe.place(x=0,y=180,width=280,height=150)
        
        patientsc_x=ttk.Scrollbar(side_leftpatientframe,orient=HORIZONTAL)
        patientsc_x.pack(side=BOTTOM,fill=X)
        patientsc_y=ttk.Scrollbar(side_leftpatientframe,orient=VERTICAL)
        patientsc_y.pack(side=RIGHT,fill=Y)
        
        self.patient_table=ttk.Treeview(side_leftpatientframe,columns=("patientid","patientname","address","symtoms","doctorid"),xscrollcommand=patientsc_x.set,yscrollcommand=patientsc_y.set)
        
        patientsc_x.config(command=self.patient_table.xview)
        patientsc_y.config(command=self.patient_table.yview)
        self.patient_table["show"]="headings"
        self.patient_table.heading("patientid",text="Patient ID")
        self.patient_table.heading("patientname",text="Patient Name")
        self.patient_table.heading("address",text="Address")
        self.patient_table.heading("symtoms",text="Symtoms")
        self.patient_table.heading("doctorid",text="Doctor ID")
        self.patient_table.pack(fill=BOTH,expand=1)

        self.patient_table.column("patientid",width=80)
        self.patient_table.column("patientname",width=80)
        self.patient_table.column("address",width=80)
        self.patient_table.column("symtoms",width=80)
        self.patient_table.column("doctorid",width=80)
        self.fetch_datapatient()
        self.patient_table.bind("<ButtonRelease-1>",self.Medget_cursorpatient)
        
        down_frame1=Frame(DataFramemid,border=4,relief=RIDGE,bg="darkgreen")
        down_frame1.place(x=285,y=185,width=80,height=140)
        
        btnAddmed=Button(down_frame1,text="ADD",font=("arial",8,"bold"),bg="lime",pady=2,fg="white",width=9,command=self.Addpatient)
        btnAddmed.grid(row=0,column=0)
        btnAddmed=Button(down_frame1,text="UPDATE",font=("arial",8,"bold"),bg="purple",pady=2,fg="white",width=9,command=self.Updatepatient)
        btnAddmed.grid(row=1,column=0)
        btnAddmed=Button(down_frame1,text="DELETE",font=("arial",8,"bold"),bg="red",pady=2,fg="white",width=9,command=self.Deletepatient)
        btnAddmed.grid(row=2,column=0)
        btnAddmed=Button(down_frame1,text="CLEAR",font=("arial",8,"bold"),bg="orange",pady=2,fg="white",width=9,command=self.clearpatient)
        btnAddmed.grid(row=3,column=0)
        btnAddmed=Button(down_frame1,text="REFRESH",font=("arial",8,"bold"),bg="green",pady=2,fg="white",width=9,command=self.refreshpatient)
        btnAddmed.grid(row=4,column=0)
        
        ###########################################  MedicineStock Table  ##############################################################
        
        
        
        #MedicineStock Variables
        self.medname_var=StringVar()
        self.refMed_var=StringVar()
        self.companyname_var=StringVar()
        self.medquantity_var=StringVar()
        self.expdate_var=StringVar()
        
        DataFrameRight=LabelFrame(DataFrame,bd=10,relief=RIDGE,padx=20,text="Medicine Stock",fg="darkgreen",font=("arial",10,"bold"))
        DataFrameRight.place(x=830,y=5,width=420,height=360)
        
        lblrefno=Label(DataFrameRight,font=("arial",12,"bold"),text="Reference No:",padx=15,pady=6)
        lblrefno.place(x=0,y=10)
        txtrefno=Entry(DataFrameRight,textvariable=self.refMed_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtrefno.place(x=145,y=10)
        
        lblmedname=Label(DataFrameRight,font=("arial",12,"bold"),text="Company Name:",padx=15,pady=6)
        lblmedname.place(x=0,y=40)
        txtmedname=Entry(DataFrameRight,textvariable=self.companyname_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtmedname.place(x=145,y=40)
        
        lblmedname=Label(DataFrameRight,font=("arial",12,"bold"),text="Medicine Name:",padx=15,pady=6)
        lblmedname.place(x=0,y=70)
        txtmedname=Entry(DataFrameRight,textvariable=self.medname_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtmedname.place(x=145,y=70)
        
        lblmedname=Label(DataFrameRight,font=("arial",12,"bold"),text="Quantity:",padx=15,pady=6)
        lblmedname.place(x=0,y=100)
        txtmedname=Entry(DataFrameRight,textvariable=self.medquantity_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtmedname.place(x=145,y=100)
        
        lblmedname=Label(DataFrameRight,font=("arial",12,"bold"),text="Exp Date:",padx=15,pady=6)
        lblmedname.place(x=0,y=130)
        txtmedname=Entry(DataFrameRight,textvariable=self.expdate_var,font=("arial",13,"bold"),bg="white",relief=RIDGE,width=24)
        txtmedname.place(x=145,y=130)
                
        side_frame=Frame(DataFrameRight,bd=4,relief=RIDGE,bg="white")
        side_frame.place(x=0,y=170,width=290,height=160)
        sct_x=ttk.Scrollbar(side_frame,orient=HORIZONTAL)
        sct_x.pack(side=BOTTOM,fill=X)
        sct_y=ttk.Scrollbar(side_frame,orient=VERTICAL)
        sct_y.pack(side=RIGHT,fill=Y)
        
        self.medicine_table=ttk.Treeview(side_frame,columns=("ref","companyname","medname","quantity","expdate"),xscrollcommand=sct_x.set,yscrollcommand=sct_y.set)
        sct_x.config(command=self.medicine_table.xview)
        sct_y.config(command=self.medicine_table.yview)
        self.medicine_table["show"]="headings"
        self.medicine_table.heading("ref",text="Ref")
        self.medicine_table.heading("companyname",text="Company Name")
        self.medicine_table.heading("medname",text="Medicine Name")
        self.medicine_table.heading("quantity",text="Quantity")
        self.medicine_table.heading("expdate",text="Exp Date")
        self.medicine_table.pack(fill=BOTH,expand=1)
        
        self.medicine_table.column("ref",width=100)
        self.medicine_table.column("medname",width=100)
        self.medicine_table.column("companyname",width=100)
        self.medicine_table.column("quantity",width=100)
        self.medicine_table.column("expdate",width=100)
        self.fetch_dataMed()
        self.medicine_table.bind("<ButtonRelease-1>",self.Medget_cursor)
        
        down_frame=Frame(DataFrameRight,border=4,relief=RIDGE,bg="darkgreen")
        down_frame.place(x=300,y=185,width=80,height=140)
        
        btnAddmed=Button(down_frame,text="ADD",font=("arial",8,"bold"),bg="lime",pady=2,fg="white",width=9,command=self.AddMed)
        btnAddmed.grid(row=0,column=0)
        btnAddmed=Button(down_frame,text="UPDATE",font=("arial",8,"bold"),bg="purple",pady=2,fg="white",width=9,command=self.UpdateMed)
        btnAddmed.grid(row=1,column=0)
        btnAddmed=Button(down_frame,text="DELETE",font=("arial",8,"bold"),bg="red",pady=2,fg="white",width=9,command=self.DeleteMed)
        btnAddmed.grid(row=2,column=0)
        btnAddmed=Button(down_frame,text="CLEAR",font=("arial",8,"bold"),bg="orange",pady=2,fg="white",width=9,command=self.clearmed)
        btnAddmed.grid(row=3,column=0)
        btnAddmed=Button(down_frame,text="REFRESH",font=("arial",8,"bold"),bg="green",pady=2,fg="white",width=9,command=self.refresh)
        btnAddmed.grid(row=4,column=0)

###################################################  Doctor Table  #############################################################

        #Doctor Table Variables
        self.doctordoctorid_var=StringVar()
        self.doctorpatientid_var=StringVar()
        self.doctormedname=StringVar()
        self.doctordoge_var=StringVar()
        self.doctorpatientname_var=StringVar()
        self.doctortotalamount=StringVar()
  
        DataFrameDoctor=LabelFrame(DataFrame,bd=10,relief=RIDGE,padx=20,text="Doctor",fg="darkgreen",font=("arial",10,"bold"))
        DataFrameDoctor.place(x=0,y=365,width=1250,height=100)
        
        lbldoctorid=Label(DataFrameDoctor,font=("arial",12,"bold"),text="Doctor ID:",padx=15,pady=6)
        lbldoctorid.grid(row=0,column=0,sticky=W)
        
        conndoctorid=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursordoctorid=conndoctorid.cursor()
        my_cursordoctorid.execute("select distinct DoctorID from Patient")
        rowsdoctorid=my_cursordoctorid.fetchall()
        ref_combodoctorid=ttk.Combobox(DataFrameDoctor,textvariable=self.doctordoctorid_var,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combodoctorid["values"]=rowsdoctorid
        ref_combodoctorid.current(0)
        ref_combodoctorid.grid(row=0,column=1)
        
        lbldoctorpatientid=Label(DataFrameDoctor,font=("arial",12,"bold"),text="Patient ID",padx=15,pady=6)
        lbldoctorpatientid.grid(row=1,column=0,sticky=W)
        
        conndoctorpatientid=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursordoctorpatientid=conndoctorpatientid.cursor()
        my_cursordoctorpatientid.execute("select PatientID from Patient")
        rowsdoctorpatientid=my_cursordoctorpatientid.fetchall()
        ref_combodoctorpatientid=ttk.Combobox(DataFrameDoctor,textvariable=self.doctorpatientid_var,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combodoctorpatientid["values"]=rowsdoctorpatientid
        ref_combodoctorpatientid.current(0)
        ref_combodoctorpatientid.grid(row=1,column=1)
        
        lbldoctormedname=Label(DataFrameDoctor,font=("arial",12,"bold"),text="Medicine Name:",padx=15,pady=6)
        lbldoctormedname.grid(row=0,column=5)
        conndoctormedname=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursordotormedname=conndoctormedname.cursor()
        my_cursordotormedname.execute("select distinct MedicineName from MedicineStock")
        rowsdoctormedname=my_cursordotormedname.fetchall()
        ref_combodoctormedname=ttk.Combobox(DataFrameDoctor,textvariable=self.doctormedname,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combodoctormedname["values"]=rowsdoctormedname
        ref_combodoctormedname.current(0)
        ref_combodoctormedname.grid(row=0,column=6)
        
        lbldoctordose=Label(DataFrameDoctor,font=("arial",12,"bold"),text="Med Dosage",padx=15,pady=6)
        lbldoctordose.grid(row=1,column=3)
        txtdoctordose=Entry(DataFrameDoctor,textvariable=self.doctordoge_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtdoctordose.grid(row=1,column=4)
        
        
        lbldoctorpatientname=Label(DataFrameDoctor,font=("arial",12,"bold"),text="Patient Name",padx=15,pady=6)
        lbldoctorpatientname.grid(row=0,column=3)
        conndoctorpatientname=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursordotorptientname=conndoctorpatientname.cursor()
        my_cursordotorptientname.execute("select distinct PatientName from Patient")
        rowsdoctorpatientname=my_cursordotorptientname.fetchall()
        ref_combodoctorpatientname=ttk.Combobox(DataFrameDoctor,textvariable=self.doctorpatientname_var,width=22,font=("arial",12,"bold"),state="readonly")
        ref_combodoctorpatientname["values"]=rowsdoctorpatientname
        ref_combodoctorpatientname.current(0)
        ref_combodoctorpatientname.grid(row=0,column=4)

        
        btnAdddoctor=Button(DataFrameDoctor,text="PRESCRIBE",font=("arial",8,"bold"),bg="GREEN",pady=2,fg="white",width=15,command=self.Adddoctor)
        btnAdddoctor.place(x=910,y=40)
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
#+++++++++++++++++++++++++++++++Billing++++++++++++++++++++++++

        self.billID_var=StringVar()
        self.billdoctorid=self.doctordoctorid_var
        self.billpatientid=self.doctorpatientid_var
        self.billpatientname_var=self.doctorpatientname_var
        self.billtotal_var=StringVar()

        DataFrameBilling=LabelFrame(DataFrame,bd=10,relief=RIDGE,padx=20,text="Billing",fg="darkgreen",font=("arial",10,"bold"))
        DataFrameBilling.place(x=0,y=465,width=1250,height=100)
        
        lblbillid=Label(DataFrameBilling,font=("arial",12,"bold"),text="Bill ID",padx=15,pady=6)
        lblbillid.grid(row=0,column=0,sticky=W)
        combillid=ttk.Combobox(DataFrameBilling,textvariable=self.billID_var,state="readonly",font=("arial",12,"bold"),width=22)
        combillid["values"]=("9001","9002","9003","9004","9005","9006","9007")
        combillid.current(0)
        combillid.grid(row=0,column=1)
        
        lblbilldoctorid=Label(DataFrameBilling,font=("arial",12,"bold"),text="Doctor ID",padx=15,pady=6)
        lblbilldoctorid.grid(row=1,column=0,sticky=W)
        txtbilldoctorid=Label(DataFrameBilling,textvariable=self.billdoctorid,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=22)
        txtbilldoctorid.grid(row=1,column=1,sticky=W)
        
        lblbillpatientid=Label(DataFrameBilling,font=("arial",12,"bold"),text="Patient ID",padx=15,pady=6)
        lblbillpatientid.grid(row=0,column=2,sticky=W)
        txtbillpatientid=Label(DataFrameBilling,textvariable=self.billpatientid,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=22)
        txtbillpatientid.grid(row=0,column=3,sticky=W)
        
        lblbillpatientname=Label(DataFrameBilling,font=("arial",12,"bold"),text="Patient Name",padx=15,pady=6)
        lblbillpatientname.grid(row=1,column=2,sticky=W)
        txtbillpatientname=Label(DataFrameBilling,textvariable=self.billpatientname_var,font=("arial",13,"bold"),bg="white",bd=2,relief=RIDGE,width=24)
        txtbillpatientname.grid(row=1,column=3,sticky=W)
        
        lblbilltotalamount=Label(DataFrameBilling,font=("arial",12,"bold"),text="Amount",padx=15,pady=6)
        lblbilltotalamount.grid(row=0,column=4,sticky=W)
        txtbilltotalamount=Entry(DataFrameBilling,textvariable=self.billtotal_var,font=("arial",12,"bold"),bg="white",relief=RIDGE,width=22)
        txtbilltotalamount.grid(row=0,column=5,sticky=W)
        
        btnbillPrint=Button(DataFrameBilling,text="Bill",font=("arial",8,"bold"),bg="red",pady=2,fg="white",width=15,command=self.Billing)
        btnbillPrint.place(x=910,y=40)
        
        
        
    def Billing(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("insert into Bill(BillID,PatientID,DoctorID,PatientName,Amount) values(%s,%s,%s,%s,%s)",(
                                                                     self.billID_var.get(),
                                                                     self.billpatientid.get(),
                                                                     self.billdoctorid.get(),   
                                                                     self.billpatientname_var.get(),
                                                                     self.billtotal_var.get()                                                              
                                                                     ))

        # DataFrame=Frame(self.root,bd=15,relief=RIDGE)
        # DataFrame.place(x=0,y=40,width=1275,height=600)
        
        output_file = 'Bill.xlsx'
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()

        my_cursor.execute("SELECT * from Bill")
        rows = my_cursor.fetchall()

        main_list = ['BillID',
                    'PatientID',
                    'DoctorID',
                    'PatientName',
                    'Amount']

        wb = Workbook()
        ws = wb.active

        # Write header row
        row_num = 1
        column_num = 1
        for field in main_list:
            ws.cell(row=row_num, column=column_num, value=field)
            column_num += 1

        # Write data rows
        row_num = 2
        for row in rows:
            column_num = 1
            for field_value in row:
                ws.cell(row=row_num, column=column_num, value=field_value)
                column_num += 1
            row_num += 1

        # Save workbook
        wb.save(output_file)

        # Close database connection
    
        conn.commit()
        conn.close()
        messagebox.showinfo("Success","Billing Done")

        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
#############################################################################3
################################################################################

############3SQL For doctor table+++++++++++++++++++
    def Adddoctor(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("insert into Doctor(DoctorID,PatientID,MedicineName,Dosage,PatientName) values(%s,%s,%s,%s,%s)",(
                                                                     self.doctordoctorid_var.get(),
                                                                     self.doctorpatientid_var.get(),
                                                                     self.doctormedname.get(),   
                                                                     self.doctordoge_var.get(),
                                                                     self.doctorpatientname_var.get()                                                              
                                                                     ))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success","Medicine Added")
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        

        
        
    #+++++++++++++++  Stock SQL Part   ++++++++++++++++++++++++++++
    def AddMed(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("insert into MedicineStock(Ref,CompanyName,MedicineName,Quantity,ExpDate) values(%s,%s,%s,%s,%s)",(
                                                                     self.refMed_var.get(),
                                                                     self.companyname_var.get(),
                                                                     self.medname_var.get(),   
                                                                     self.medquantity_var.get(),
                                                                     self.expdate_var.get()                                                                 
                                                                     ))
        conn.commit()
        self.fetch_dataMed()
        self.Medget_cursor()
        conn.close()
        messagebox.showinfo("Success","Medicine Added")
        
    def refresh(self):
        self.fetch_dataMed()
    
    def fetch_dataMed(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("select * from MedicineStock")
        rows=my_cursor.fetchall()
        if len(rows)!=0:
            self.medicine_table.delete(*self.medicine_table.get_children())
            for i in rows:
                self.medicine_table.insert("",END,values=i)
            conn.commit()
        conn.close()
        
        
    def Medget_cursor(self,event=""):
        cursor_row=self.medicine_table.focus()
        content=self.medicine_table.item(cursor_row)
        row=content["values"]
        self.refMed_var.set(row[0])
        self.companyname_var.set(row[1])
        self.medname_var.set(row[2])
        self.medquantity_var.set(row[3])
        self.expdate_var.set(row[4])

        
    def UpdateMed(self):
        if self.refMed_var.get()=="" or self.medname_var.get()=="" or self.medquantity_var.get()=="":
            messagebox.showerror("Error","All fields are Required!")
        else:
            conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
            my_cursor=conn.cursor()
            my_cursor.execute("update MedicineStock set CompanyName=%s,MedicineName=%s,Quantity=%s ,ExpDate=%s where Ref=%s",(
                                                                        self.companyname_var.get(),
                                                                        self.medname_var.get(),
                                                                        self.medquantity_var.get(),
                                                                        self.expdate_var.get(),
                                                                        self.refMed_var.get()                                                                     
                                                                        ))
            conn.commit()
            self.fetch_dataMed()
            conn.close()
            messagebox.showinfo("Success","Medicine has been updated")
        
        
    def DeleteMed(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        sql="delete from MedicineStock where Ref=%s"
        val=(self.refMed_var.get(),)
        my_cursor.execute(sql,val)
        conn.commit()
        self.fetch_dataMed()
        self.fetch_dataMedinfo()
        conn.close()
        
    def clearmed(self):
        self.refMed_var.set("")
        self.medname_var.set("")
        self.companyname_var.set("")
        self.medquantity_var.set("")
        self.expdate_var.set("")
        
#_______________________________________________





 #^^^^^^^^^^^^^^^^^^Now SQL Part For Left Data Frame*^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^        
    def AddMedinfo(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("insert into MedicineInfo(StoreID,Ref,MedType,SideEffect,Price) values(%s,%s,%s,%s,%s)",(
                                                                     self.storeid_var.get(),
                                                                     self.refinfo_var.get(),
                                                                     self.typemed_var.get(),   
                                                                     self.sideeffect_var.get(),
                                                                     self.price_var.get(),                                                              
                                                                     ))
        conn.commit()
        self.fetch_dataMedinfo()
        self.Medget_cursorinfo()
        conn.close()
        messagebox.showinfo("Success","Medicine Info Added")
        
    def refreshinfo(self):
        self.fetch_dataMedinfo()
    
    def fetch_dataMedinfo(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("select * from MedicineInfo")
        rows=my_cursor.fetchall()
        if len(rows)!=0:
            self.medicineinfo_table.delete(*self.medicineinfo_table.get_children())
            for i in rows:
                self.medicineinfo_table.insert("",END,values=i)
            conn.commit()
        conn.close()
        
        
    def Medget_cursorinfo(self,event=""):
        cursor_row=self.medicineinfo_table.focus()
        content=self.medicineinfo_table.item(cursor_row)
        row=content["values"]
        self.storeid_var.set(row[0]),
        self.refinfo_var.set(row[1]),
        self.typemed_var.set(row[2]),   
        self.sideeffect_var.set(row[3]),
        self.price_var.set(row[4])

        
    def UpdateMedinfo(self):
        if self.storeid_var.get()=="" or self.typemed_var.get()=="":
            messagebox.showerror("Error","All fields are Required!")
        else:
            conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
            my_cursor=conn.cursor()
            my_cursor.execute("update MedicineInfo set MedType=%s,SideEffect=%s,Price=%s where Ref=%s",(

                                                                        self.typemed_var.get(),
                                                                        self.sideeffect_var.get(),
                                                                        self.price_var.get(),
                                                                        self.refinfo_var.get()                                                                     
                                                                        ))
            conn.commit()
            self.fetch_dataMedinfo()
            conn.close()
            messagebox.showinfo("Success","Medicine has been updated")
        
        
    def DeleteMedinfo(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        sql="delete from MedicineInfo where Ref=%s"
        val=(self.refinfo_var.get(),)
        my_cursor.execute(sql,val)
        conn.commit()
        self.fetch_dataMedinfo()
        conn.close()
        
    def clearmedinfo(self):
        self.storeid_var.set("")
        self.typemed_var.set("")
        self.sideeffect_var.set("")
        self.price_var.set("")     
   #+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   
   
   
   
   
   
   #++++++++++++++++++++Patient Sql Part+++++++++++++++++++++++++
   
   
    def Addpatient(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("insert into Patient(PatientID,PatientName,Address,Symtoms,DoctorID) values(%s,%s,%s,%s,%s)",(
                                                                     self.patientid_var.get(),
                                                                     self.patientnamevar.get(),
                                                                     self.addresspatient_var.get(),   
                                                                     self.symptoms_var.get(),
                                                                     self.patientdoctorid_var.get()                                                                 
                                                                     ))
        conn.commit()
        self.fetch_datapatient()
        self.Medget_cursorpatient()
        conn.close()
        messagebox.showinfo("Success","Patient Added")
        self.root()
        
    def refreshpatient(self):
        self.fetch_datapatient()
    
    def fetch_datapatient(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        my_cursor.execute("select * from Patient")
        rows=my_cursor.fetchall()
        if len(rows)!=0:
            self.patient_table.delete(*self.patient_table.get_children())
            for i in rows:
                self.patient_table.insert("",END,values=i)
            conn.commit()
        conn.close()
        
        
    def Medget_cursorpatient(self,event=""):
        cursor_row=self.patient_table.focus()
        content=self.patient_table.item(cursor_row)
        row=content["values"]
        self.patientid_var.set(row[0])
        self.patientnamevar.set(row[1])
        self.addresspatient_var.set(row[2])
        self.symptoms_var.set(row[3])
        self.patientdoctorid_var.set(row[4])

        
    def Updatepatient(self):
        if self.patientnamevar.get()=="" or self.addresspatient_var.get()=="" or self.symptoms_var.get()=="":
            messagebox.showerror("Error","All fields are Required!")
        else:
            conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
            my_cursor=conn.cursor()
            my_cursor.execute("update Patient set PatientName=%s,Address=%s,Symtoms=%s where PatientID=%s",(
                                                                        self.patientnamevar.get(),
                                                                        self.addresspatient_var.get(),
                                                                        self.symptoms_var.get(),
                                                                        self.patientid_var.get()                                                                     
                                                                        ))
            conn.commit()
            self.fetch_datapatient()
            conn.close()
            messagebox.showinfo("Success","Patient Data has been updated")
        
        
    def Deletepatient(self):
        conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
        my_cursor=conn.cursor()
        sql="delete from Patient where PatientID=%s"
        val=(self.patientid_var.get(),)
        my_cursor.execute(sql,val)
        conn.commit()
        self.fetch_datapatient()
        conn.close()
        
    def clearpatient(self):
        self.patientnamevar.set("")
        self.addresspatient_var.set("")
        self.symptoms_var.set("")
        self.patientdoctorid_var.set("")
   
   

   
   
   
   
   
   
   
   
   
   
   
   
   
   
   
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    #     #SQL BACKENd
    #     #Add Medicine Functionalityvdecleration
        

        
    
        
    # #{{{{{{{{{{{{{{{{{{{{{{{{Main table}}}}}}}}}}}}}}}}}}}}}}}}
        
    # def Add_data(self):
    #     if self.ref_var.get()=="" or self.lot_var.get()=="":
    #         messagebox.showerror("Error","All Fields Required")
    #     else:
    #         conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
    #         my_cursor=conn.cursor()
    #         my_cursor.execute("insert into pharmacy values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(
    #                                                                             self.ref_var.get(),
    #                                                                             self.cmpName_var.get(),
    #                                                                             self.typemed_var.get(),
    #                                                                             self.medName_var.get(),
    #                                                                             self.lot_var.get(),
    #                                                                             self.issuedate_var.get(),
    #                                                                             self.expdate_var.get(),
    #                                                                             self.uses_var.get(),
    #                                                                             self.sideeffect_var.get(),
    #                                                                             self.warning_var.get(),
    #                                                                             self.dosage_var.get(),
    #                                                                             self.price_var.get(),
    #                                                                             self.product_var.get()
                                                                        
    #                                                                     ))
    #         conn.commit()
    #         self.fetch_data()
    #         conn.close()
    #         messagebox.showinfo("Success","Data has been Inserted")
   
    # def fetch_data(self):
    #     conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
    #     my_cursor=conn.cursor()
    #     my_cursor.execute("select * from pharmacy")
    #     row=my_cursor.fetchall()
    #     if len(row)!=0:
    #         self.pharmacy_table.delete(*self.pharmacy_table.get_children())
    #         for i in row:
    #             self.pharmacy_table.insert("",END,values=i)
    #         conn.commit()
    #     conn.close()
            
    # def get_cursor(self,ev=""):
    #     cursor_row=self.pharmacy_table.focus()
    #     content=self.pharmacy_table.item(cursor_row)
    #     row=content["values"]
        
    #     self.ref_var.set(row[0]),
    #     self.cmpName_var.set(row[1]),
    #     self.typemed_var.set(row[2]),
    #     self.medName_var.set(row[3]),
    #     self.lot_var.set(row[4]),
    #     self.issuedate_var.set(row[5]),
    #     self.expdate_var.set(row[6]),
    #     self.uses_var.set(row[7]),
    #     self.sideeffect_var.set(row[8]),
    #     self.warning_var.set(row[9]),
    #     self.dosage_var.set(row[10]),
    #     self.price_var.set(row[11]),
    #     self.product_var.set(row[12])
            
    # def Update(self):
    #     if self.ref_var.get()=="" or self.lot_var.get()=="":
    #         messagebox.showerror("Error","All fields are Required!")
    #     else:
    #         conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
    #         my_cursor=conn.cursor()
    #         my_cursor.execute("update pharmacy set CmpName=%s,TypeMed=%s,medname=%s,LotNo=%s,Issuedate=%s,Expdate=%s,uses=%s,Sideeffect=%s,warning=%s,dosage=%s,Price=%s,product=%s where Ref_no=%s",(
    #                                                                     self.cmpName_var.get(),
    #                                                                     self.typemed_var.get(),
    #                                                                     self.medName_var.get(),
    #                                                                     self.lot_var.get(),
    #                                                                     self.issuedate_var.get(),
    #                                                                     self.expdate_var.get(),
    #                                                                     self.uses_var.get(),
    #                                                                     self.sideeffect_var.get(),
    #                                                                     self.warning_var.get(),
    #                                                                     self.dosage_var.get(),
    #                                                                     self.price_var.get(),
    #                                                                     self.product_var.get(),
    #                                                                     self.ref_var.get()
    #                                                                     ))
    #         conn.commit()
    #         self.fetch_data()
    #         conn.close()
    #         messagebox.showinfo("Success","Record has been updated Successfully") 
            
    
    
    # def delete(self):
    #     conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
    #     my_cursor=conn.cursor()
    #     sql="delete from pharmacy where Ref_no=%s"
    #     val=(self.ref_var.get(),)
    #     my_cursor.execute(sql,val)
    #     conn.commit()
    #     self.fetch_data()
    #     conn.close()
    #     messagebox.showinfo("DELETED","Info Deleted Successfully")
        
        
    # def reset(self):
    #     #self.ref_var.set(""),
    #     self.cmpName_var.set(""),
    #     #self.typemed_var.set(""),
    #     #self.medName_var.set(""),
    #     self.lot_var.set(""),
    #     self.issuedate_var.set(""),
    #     self.expdate_var.set(""),
    #     self.uses_var.set(""),
    #     self.sideeffect_var.set(""),
    #     self.warning_var.set(""),
    #     self.dosage_var.set(""),
    #     self.price_var.set(""),
    #     self.product_var.set("")
        
    # def search_data(self):
    #     conn=mysql.connector.connect(host="localhost",username="root",password="Prishabh@7044",database="mydata")
    #     my_cursor=conn.cursor()
    #     my_cursor.execute("select * from pharmacy where "+str(self.search_var.get())+"LIKE"+str(self.searchTxt_var.get())+"%")
        
    #     roww=my_cursor.fetchall()
    #     if len(roww)!=0:
    #         self.pharmacy_table.delete(*self.pharmacy_table.get_children())
    #         for i in roww:
    #             self.pharmacy_table.insert("",END,values=i)
    #             conn.commit()
    #         conn.close
        
        
        
        
        
        
        
        
        
        
        
        
        
        
if __name__ == "__main__":
    root=Tk()
    obj=PharmacyManagementSystem(root)
    root.mainloop()